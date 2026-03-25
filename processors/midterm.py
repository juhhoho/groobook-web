from io import BytesIO
import openpyxl


def process_midterm(
    file_bytes: bytes,
    main_banding: int = 20,
    mini_banding: int = 70,
    teacher_banding: int = 15,
    progress_cb=None,
) -> list:
    """
    중간기말교재 Excel 분석 → slip list 반환

    중간기말교재 구조:
    - 책(출판사)이 열(가로)에 위치
    - 캠퍼스가 행(세로)에 위치
    - 짝수행 = TR(교사용) 수량, 홀수행 = 전체; 학생용 = 전체 - TR

    slip 구조 (8요소):
    [캠퍼스, 학년+시험명, "N번", 출판사, "N부", 총덩이수, "-", 현재번호]
    정렬: 본책 전체 → 미니북 전체 → 교사용 전체
          각 타입 내: 책 열 인덱스 오름차순 → 캠퍼스 행 인덱스 오름차순
    """
    def cb(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)

    cb(10, "파일 읽는 중...")
    workbook = openpyxl.load_workbook(BytesIO(file_bytes))
    sheet1 = workbook.worksheets[0]

    cb(25, "데이터 추출 중...")

    # A2: 학년 + 시험 정보
    cell_value = sheet1['A2'].value
    words = cell_value.split()
    quarter_info = " ".join(words[1:3])
    grade = " ".join(words[-1:])

    # Row 3 (C열~"계"전): 출판사 목록
    publishers = []
    for cell in sheet1[3][2:-1]:
        if cell.value == "계":
            break
        else:
            publishers.append(cell.value)

    # A열 (row 4~, 2행씩, "총계" 전): 캠퍼스 목록
    campuses = []
    row_number = 4
    while True:
        cell_value = sheet1.cell(row=row_number, column=1).value
        if str(cell_value)[-2:] == "총계":
            break
        campuses.append(cell_value)
        row_number += 2

    # 마지막 열("계" 열) 찾기
    last_column = None
    for col in range(1, sheet1.max_column + 1):
        if sheet1.cell(row=3, column=col).value == "계":
            last_column = col
            break

    cb(40, "수량 분리 중...")

    # outlist1: TR(교사용) 수량 [출판사][캠퍼스]
    # outlist2: 학생용 수량 = 전체 - TR [출판사][캠퍼스]
    outlist1 = []
    outlist2 = []

    for col in range(3, last_column):
        inlist1 = []  # TR (짝수행)
        inlist2 = []  # 전체 (홀수행)
        row = 4
        for _ in range(4, 4 + 2 * len(campuses)):
            cell_value = sheet1.cell(row=row, column=col).value
            if type(cell_value) != int:
                break
            if row % 2 == 0:
                inlist1.append(cell_value)
            else:
                inlist2.append(cell_value)
            row += 1

        if len(inlist1) != 0:
            outlist1.append(inlist1)
        if len(inlist2) != 0:
            # 학생용 = 전체 - TR
            inlist3 = [inlist2[k] - inlist1[k] for k in range(len(campuses))]
            outlist2.append(inlist3)

    cb(55, "포장 계산 중...")

    def make_slips(data_matrix, banding, type_label):
        """
        data_matrix: [책 index][캠퍼스 index] → 수량
        slip 구조 (6섹션 레이아웃):
        [캠퍼스이름, 학년+타입, 캠퍼스인덱스(1부터), 출판사, 부수, 총덩이수, "-", 현재번호]
        """
        slips = []
        for pub_idx, data_row in enumerate(data_matrix):
            pub_name = publishers[pub_idx]
            pub_num = pub_idx + 1  # 출판사 인덱스 1부터 시작
            for cam_idx, quantity in enumerate(data_row):
                campus = campuses[cam_idx]
                if not quantity or quantity == 0:
                    continue
                packing = quantity // banding + (1 if quantity % banding != 0 else 0)
                remaining = quantity
                for counter in range(1, packing + 1):
                    qty_str = f"{banding}부" if remaining - banding > 0 else f"{remaining}부"
                    slips.append([campus, f"{grade} {type_label}", pub_num, pub_name, qty_str, packing, "-", counter])
                    if remaining - banding > 0:
                        remaining -= banding
        return slips

    # 본책 → 미니북 → 교사용 순서
    slip_list = []
    slip_list.extend(make_slips(outlist2, main_banding, "본책"))
    slip_list.extend(make_slips(outlist2, mini_banding, "미니북"))
    slip_list.extend(make_slips(outlist1, teacher_banding, "교사용"))

    cb(75, "완료 준비 중...")
    workbook.close()
    return slip_list
