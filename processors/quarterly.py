from io import BytesIO
import openpyxl


def process_quarterly(file_bytes: bytes, progress_cb=None) -> list:
    """
    분기별교재 Excel 분석 → slip list 반환

    slip 구조 (8요소):
    [캠퍼스, 분기정보, "N 번", 교재명, "N부", 총덩이수, "_", 현재번호]
    정렬: 책 행 인덱스 오름차순 → 캠퍼스 열 인덱스 오름차순 (iterate 순서 유지)
    """
    def cb(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)

    cb(10, "파일 읽는 중...")
    workbook = openpyxl.load_workbook(BytesIO(file_bytes))
    sheet1 = workbook.worksheets[0]

    cb(30, "데이터 추출 중...")

    # A1: 분기 정보 (앞 두 단어)
    cell_value = sheet1['A1'].value
    words = cell_value.split()
    quarter_info = " ".join(words[:2])

    last_row = sheet1.max_row
    last_column = sheet1.max_column
    campus_col_start = 3  # C열

    # "밴딩" 열 찾기
    for col in range(campus_col_start, last_column + 1):
        if sheet1.cell(row=2, column=col).value == "밴딩":
            last_column = col
            break

    cb(55, "포장 계산 중...")

    slip_list = []
    index = 1

    # Row 3+: 책 행 순서 → 캠퍼스 열 순서
    for row in range(3, last_row):
        book_name = sheet1.cell(row=row, column=2).value
        col = campus_col_start

        while col <= last_column - 2:
            campus_name = sheet1.cell(row=2, column=col).value

            if campus_name in ["계", "밴딩", "옵셋"]:
                col += 2
                continue
            if not campus_name:
                break

            # 밴딩값
            banding = sheet1.cell(row=row, column=last_column).value
            try:
                if banding is not None:
                    banding = int(banding)
            except ValueError:
                raise

            # 수량
            quantity = sheet1.cell(row=row, column=col).value
            try:
                quantity = 0 if quantity is None else int(quantity)
            except ValueError:
                raise

            if quantity == 0:
                packing = 0
            else:
                packing = quantity // banding + (1 if quantity % banding != 0 else 0)

            # 밴딩 분할 slip 생성
            counter = 1
            remaining = quantity
            while counter <= packing:
                if remaining - banding > 0:
                    slip_list.append([campus_name, quarter_info, index, book_name, f"{banding}부", packing, "_", counter])
                    remaining -= banding
                else:
                    slip_list.append([campus_name, quarter_info, index, book_name, f"{remaining}부", packing, "_", counter])
                counter += 1

            col += 1
        index += 1

    cb(75, "완료 준비 중...")
    workbook.close()
    return slip_list
