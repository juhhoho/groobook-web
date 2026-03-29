import re
from io import BytesIO

import openpyxl
from openpyxl.utils import column_index_from_string

CAMPUS_FULL = [
    "대치본원", "대치본원S관", "송파캠퍼스", "중계캠퍼스", "분당캠퍼스",
    "수지캠퍼스", "수지3관캠퍼스", "죽전캠퍼스", "1동탄캠퍼스", "2동탄캠퍼스",
    "영통캠퍼스", "수원캠퍼스", "평촌캠퍼스", "평촌고등관", "일산캠퍼스",
    "강동하남캠퍼스", "SIS",
]

CAMPUS_SHORT = [
    "대치", "대치S관", "송파", "중계", "분당",
    "수지", "수지3관", "죽전", "1동탄", "2동탄",
    "영통", "수원", "평촌", "평촌고등관", "일산",
    "강동하남", "SIS",
]

TYPE_NOTE, TYPE_BOOK, TYPE_GICHUL, TYPE_YESANG = 0, 1, 2, 3

# 발주서 종류별 (합계 키 열, 수량 열, 공급액 열) — 1-based
_COL_MAP = {
    TYPE_NOTE:   (1, 3, 6),
    TYPE_BOOK:   (2, 6, 9),
    TYPE_GICHUL: (2, 5, 8),
    TYPE_YESANG: (2, 6, 9),
}

# 정규식: 캠퍼스 단축명을 길이 내림차순으로 정렬해 longest-match 보장
_sorted_short = sorted(CAMPUS_SHORT, key=len, reverse=True)
_campus_pat = "|".join(re.escape(c) for c in _sorted_short)

_RE_NOTE   = re.compile(rf"^({_campus_pat})노트$")
_RE_BOOK   = re.compile(rf"^({_campus_pat})(중[123])(_[0-9]+)?$")
_RE_GICHUL = re.compile(rf"^({_campus_pat})기출$")
_RE_YESANG = re.compile(rf"^({_campus_pat})(중[123]예상문제)(_[0-9]+)?$")

# ──────────────── 수식 평가기 ────────────────

_RE_CELL_REF = re.compile(r'^\$?([A-Za-z]+)\$?(\d+)$')


def _cell_ref_to_rc(ref: str):
    """'$A$1' / 'A1' → (row, col) 1-based. 실패 시 (None, None)."""
    m = _RE_CELL_REF.match(ref.strip())
    if not m:
        return None, None
    return int(m.group(2)), column_index_from_string(m.group(1).upper())


def _get_val(ws, row, col, depth=0):
    """셀 값을 반환. 수식이면 재귀 평가(최대 4단계). 계산 불가 시 None."""
    if depth > 4:
        return None
    v = ws.cell(row=row, column=col).value
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str) or not v.startswith('='):
        return None
    return _eval(ws, v[1:].strip(), depth)


def _eval(ws, expr: str, depth: int):
    """간단한 Excel 수식 문자열 평가 (SUM / 사칙연산 지원)."""
    expr = expr.strip()

    # ROUND(SUM(...), n) 또는 SUM(...)
    m = re.fullmatch(r'(?:ROUND\()?SUM\(([^)]+)\)(?:,\s*-?\d+\))?', expr, re.IGNORECASE)
    if m:
        return _eval_sum(ws, m.group(1).strip(), depth)

    # 셀 참조들의 사칙연산 체인: A1+B1, D10+E10, F10*H10, A1+B1+C1 등
    # 셀 참조 패턴: 선택적 $ + 열문자 + 선택적 $ + 행숫자
    _CELL = r'\$?[A-Za-z]+\$?\d+'
    if re.fullmatch(rf'({_CELL})([+\-*/]({_CELL}))+', expr):
        parts = re.split(r'([+\-*/])', expr)
        # parts = [셀, 연산자, 셀, 연산자, 셀, ...]
        r, c = _cell_ref_to_rc(parts[0])
        if r is None:
            return None
        result = _get_val(ws, r, c, depth + 1)
        if result is None:
            return None
        for i in range(1, len(parts) - 1, 2):
            op = parts[i]
            r, c = _cell_ref_to_rc(parts[i + 1])
            if r is None:
                return None
            v = _get_val(ws, r, c, depth + 1)
            if v is None:
                return None
            if op == '+':   result += v
            elif op == '-': result -= v
            elif op == '*': result *= v
            elif op == '/' and v != 0: result /= v
            else: return None
        return result

    return None


def _eval_sum(ws, range_str: str, depth: int):
    """SUM 인자 문자열(A1:B10 또는 A1) 평가."""
    if ':' in range_str:
        parts = range_str.split(':')
        r1, c1 = _cell_ref_to_rc(parts[0])
        r2, c2 = _cell_ref_to_rc(parts[1])
        if None in (r1, c1, r2, c2):
            return None
        total = 0.0
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                v = _get_val(ws, r, c, depth + 1)
                if v is not None:
                    total += v
        return total
    else:
        r, c = _cell_ref_to_rc(range_str)
        return _get_val(ws, r, c, depth + 1) if r else None


# ──────────────────────────────────────────────


def parse_sheet_name(name: str) -> dict | None:
    """시트명을 파싱해 캠퍼스·발주서 종류·학년·추가발주 정보를 반환. 인식 불가면 None."""
    name = name.strip()
    m = _RE_NOTE.match(name)
    if m:
        return {
            "campus_idx": CAMPUS_SHORT.index(m.group(1)),
            "type_idx": TYPE_NOTE,
            "grade": 0,
            "extra": 0,
        }

    m = _RE_BOOK.match(name)
    if m:
        return {
            "campus_idx": CAMPUS_SHORT.index(m.group(1)),
            "type_idx": TYPE_BOOK,
            "grade": int(m.group(2)[1]),
            "extra": int(m.group(3)[1:]) if m.group(3) else 0,
        }

    m = _RE_GICHUL.match(name)
    if m:
        return {
            "campus_idx": CAMPUS_SHORT.index(m.group(1)),
            "type_idx": TYPE_GICHUL,
            "grade": 0,
            "extra": 0,
        }

    m = _RE_YESANG.match(name)
    if m:
        return {
            "campus_idx": CAMPUS_SHORT.index(m.group(1)),
            "type_idx": TYPE_YESANG,
            "grade": int(m.group(2)[1]),
            "extra": int(m.group(3)[1:]) if m.group(3) else 0,
        }

    return None


def _find_hapgye_row(sheet, key_col: int) -> int | None:
    """key_col 열에서 "합계" 문자열이 있는 행 번호(1-based) 반환."""
    for row in sheet.iter_rows(min_col=key_col, max_col=key_col):
        cell = row[0]
        if cell.value is not None and str(cell.value).strip() == "합계":
            return cell.row
    return None


def _get_val_with_cache(ws_data, ws_formula, row, col, depth=0):
    """캐시값 우선, 없으면 수식 평가. ws_data=data_only 시트, ws_formula=수식 시트."""
    # 1순위: data_only 캐시
    v = ws_data.cell(row=row, column=col).value
    if isinstance(v, (int, float)):
        return float(v)
    # 2순위: 수식 평가 (외부 참조는 불가, 내부 수식만 처리)
    return _get_val(ws_formula, row, col, depth)


def _extract_order_data(ws_data, ws_formula, type_idx: int) -> tuple[int, int] | None:
    """(수량, 공급액) 반환.
    ws_data: data_only=True 시트 (캐시값), ws_formula: 수식 시트 (평가 폴백)."""
    key_col, qty_col, amt_col = _COL_MAP[type_idx]
    row_num = _find_hapgye_row(ws_data, key_col)
    if row_num is None:
        row_num = _find_hapgye_row(ws_formula, key_col)
    if row_num is None:
        return None

    qty = _get_val_with_cache(ws_data, ws_formula, row_num, qty_col)
    amt = _get_val_with_cache(ws_data, ws_formula, row_num, amt_col)

    # 합계 행 평가 실패 시 위 행들을 직접 합산
    if qty is None:
        qty = sum(
            v for r in range(1, row_num)
            if (v := _get_val_with_cache(ws_data, ws_formula, r, qty_col)) is not None
        )
    if amt is None:
        amt = sum(
            v for r in range(1, row_num)
            if (v := _get_val_with_cache(ws_data, ws_formula, r, amt_col)) is not None
        )

    return (int(qty), int(amt))


def process_jundeung_bill(file_bytes: bytes, progress_cb=None) -> bytes:
    """
    청구서 엑셀 파일을 처리해 견적서를 자동 작성하고 시트를 정렬한 뒤
    수정된 엑셀 파일을 bytes로 반환.
    """
    def cb(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)

    cb(5, "파일 읽는 중...")
    # data_only=True: Excel이 캐시한 값 읽기 (외부 참조 포함)
    wb_data = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    # 수식 보존 + 정렬/저장용
    wb = openpyxl.load_workbook(BytesIO(file_bytes))

    gyeonjeok = wb.worksheets[0]
    order_sheets = wb.worksheets[1:]
    order_sheets_data = wb_data.worksheets[1:]

    cb(15, "발주서 분석 중...")

    campus_totals: dict[int, dict] = {}
    sheet_meta: list[tuple] = []

    total = len(order_sheets)
    for i, (sheet, sheet_data) in enumerate(zip(order_sheets, order_sheets_data)):
        if total > 0:
            cb(15 + int(30 * i / total), f"발주서 분석 중... ({i + 1}/{total})")

        meta = parse_sheet_name(sheet.title)
        sheet_meta.append((sheet, meta))

        if meta is None:
            continue

        data = _extract_order_data(sheet_data, sheet, meta["type_idx"])
        if data is None:
            continue

        qty, amt = data
        cidx = meta["campus_idx"]
        if cidx not in campus_totals:
            campus_totals[cidx] = {"qty": 0, "amt": 0}
        campus_totals[cidx]["qty"] += qty
        campus_totals[cidx]["amt"] += amt

    cb(50, "견적서 작성 중...")

    # 견적서 13~30행에 캠퍼스별 합계 기록
    # 열B(2)=납품처(풀네임), 열G(7)=수량, 열K(11)=금액
    write_row = 13
    for cidx, full_name in enumerate(CAMPUS_FULL):
        if cidx not in campus_totals:
            continue
        totals = campus_totals[cidx]
        gyeonjeok.cell(row=write_row, column=2).value = full_name
        gyeonjeok.cell(row=write_row, column=7).value = totals["qty"]
        gyeonjeok.cell(row=write_row, column=11).value = totals["amt"]
        write_row += 1

    cb(70, "시트 정렬 중...")

    def sort_key(item):
        _, meta = item
        if meta is None:
            return (len(CAMPUS_SHORT), 99, 99, 99)
        return (meta["campus_idx"], meta["type_idx"], meta["grade"], meta["extra"])

    sorted_order = sorted(sheet_meta, key=sort_key)
    wb._sheets = [gyeonjeok] + [s for s, _ in sorted_order]

    cb(90, "파일 저장 중...")
    output = BytesIO()
    wb.save(output)
    wb.close()

    cb(100, "완료!")
    return output.getvalue()
